import openpyxl
import pandas as pd
import os
import shutil
import time
from typing import Optional


def convert_xls_to_xlsx(xls_path: str, overwrite: bool = False) -> str:
    """
    Convert file .xls (định dạng cũ) sang .xlsx bằng Excel COM (pywin32).
    Giữ nguyên format/chart/ô gộp. Yêu cầu máy phải cài Microsoft Excel.

    Args:
        xls_path: đường dẫn file .xls
        overwrite: nếu file .xlsx đích đã tồn tại thì có ghi đè không.
                   Nếu False và file đã tồn tại -> trả về luôn đường dẫn đó (không convert lại).

    Returns:
        Đường dẫn file .xlsx (cùng thư mục, cùng tên).

    Raises:
        FileNotFoundError: không tìm thấy file .xls.
        RuntimeError: thiếu pywin32 hoặc không khởi động được Excel.
    """
    xls_path = os.path.abspath(xls_path)
    if not os.path.exists(xls_path):
        raise FileNotFoundError(f"Không tìm thấy file: {xls_path}")

    xlsx_path = os.path.splitext(xls_path)[0] + '.xlsx'
    if os.path.exists(xlsx_path) and not overwrite:
        return xlsx_path

    try:
        import win32com.client as win32
        import pythoncom
    except ImportError:
        raise RuntimeError(
            "Chưa cài 'pywin32' nên không convert được .xls. "
            "Cài bằng: pip install pywin32"
        )

    # App xử lý folder trong Thread phụ. Dùng COM (Excel) trong thread BẮT BUỘC phải
    # CoInitialize, nếu không Excel COM chập chờn ("Open method failed" lúc được lúc không).
    pythoncom.CoInitialize()
    last_err = None
    try:
        # Excel cold-start / đang bận có thể fail lần đầu -> thử lại tối đa 3 lần.
        for attempt in range(3):
            excel = None
            wb = None
            try:
                # DispatchEx -> instance Excel riêng, không chiếm Excel người dùng đang mở
                excel = win32.DispatchEx('Excel.Application')
                excel.Visible = False
                excel.DisplayAlerts = False
                try:
                    # msoAutomationSecurityForceDisable (3): CHẶN mọi macro khi mở file.
                    # Quan trọng vì file .xls cũ có thể chứa macro virus (vd X97M/Laroux).
                    excel.AutomationSecurity = 3
                except Exception:
                    pass
                # Mở ReadOnly + UpdateLinks=0 (tham số POSITIONAL — pywin32 late-binding
                # không nhận keyword args cho method này) -> tránh xung đột khi file đang mở.
                # Lưu ra .xlsx (FileFormat=51) sẽ LOẠI BỎ macro -> file kết quả sạch virus.
                wb = excel.Workbooks.Open(xls_path, 0, True)  # Filename, UpdateLinks=0, ReadOnly=True
                if os.path.exists(xlsx_path):
                    os.remove(xlsx_path)
                wb.SaveAs(xlsx_path, FileFormat=51)  # 51 = xlOpenXMLWorkbook (.xlsx)
                return xlsx_path
            except Exception as e:
                last_err = e
                time.sleep(1.5)  # chờ Excel sẵn sàng rồi thử lại
            finally:
                if wb is not None:
                    try:
                        wb.Close(False)
                    except Exception:
                        pass
                if excel is not None:
                    try:
                        excel.Quit()
                    except Exception:
                        pass
        raise RuntimeError(
            f"Không convert được .xls sang .xlsx sau 3 lần thử (cần Microsoft Excel): {last_err}"
        )
    finally:
        pythoncom.CoUninitialize()



def load_cc_per_old(file_path: str = 'CC_tuoi.xlsx') -> pd.DataFrame:

    df = pd.read_excel(file_path, skiprows=1)

    cols = ['thang_tuoi', 'minus_3sd', 'minus_2sd', 'trung_vi', 'plus_2sd', 'plus_3sd']

    df_trai = df.iloc[:, 0:6].copy()
    df_trai.columns = cols
    df_trai['gioi_tinh'] = 'trai'

    df_gai = df.iloc[:, 7:13].copy()
    df_gai.columns = cols
    df_gai['gioi_tinh'] = 'gai'

    result = pd.concat([df_trai, df_gai], ignore_index=True)
    result = result.dropna()

    return result


def load_cn_per_old(file_path: str = 'CN_tuoi.xlsx') -> pd.DataFrame:

    df = pd.read_excel(file_path, skiprows=1)

    cols = ['thang_tuoi', 'minus_3sd', 'minus_2sd', 'trung_vi', 'plus_2sd', 'plus_3sd']

    df_trai = df.iloc[:, 0:6].copy()
    df_trai.columns = cols
    df_trai['gioi_tinh'] = 'trai'

    df_gai = df.iloc[:, 7:13].copy()
    df_gai.columns = cols
    df_gai['gioi_tinh'] = 'gai'

    result = pd.concat([df_trai, df_gai], ignore_index=True)
    result = result.dropna()

    return result


def load_cn_cc(file_path: str) -> pd.DataFrame:
    """Load cân nặng theo chiều cao"""
    df = pd.read_excel(file_path, skiprows=1)

    cols = ['chieu_cao', 'minus_3sd', 'minus_2sd', 'trung_vi', 'plus_2sd', 'plus_3sd']

    df_trai = df.iloc[:, 0:6].copy()
    df_trai.columns = cols
    df_trai['gioi_tinh'] = 'trai'

    df_gai = df.iloc[:, 7:13].copy()
    df_gai.columns = cols
    df_gai['gioi_tinh'] = 'gai'

    result = pd.concat([df_trai, df_gai], ignore_index=True)
    result = result.dropna()

    return result


def load_danh_sach_can_do(file_path: str, sheet_name: str = 'Sheet1') -> pd.DataFrame:
    """Load danh sách cân đo trẻ em (format mới 2025)"""
    df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=6, header=None)
    
    # Cột: 0-STT, 1-Họ tên, 2-Nam, 3-Nữ, 4-Ngày sinh, 5-Tháng tuổi, 6-Họ tên mẹ, 
    #      7-Cân nặng, 8-TT CN/Tuổi, 9-Chiều cao, 10-TT CC/Tuổi, 11-CN/CC
    df = df.iloc[:, [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11]]
    
    df.columns = [
        'stt', 'ho_ten', 'nam', 'nu', 'ngay_sinh', 'thang_tuoi',
        'ho_ten_me', 'can_nang', 'execute_cn_tuoi', 'chieu_cao', 'execute_cc_tuoi', 'execute_cn_cc'
    ]
    
    # Loại bỏ các dòng không phải data (stt không phải số)
    stt_raw = df['stt']
    stt_numeric = pd.to_numeric(stt_raw, errors='coerce')
    # Giữ lại hàng stt là số hoặc trống/NaN; loại các ô chứa chữ/không chuyển được
    mask_keep = stt_numeric.notna() | stt_raw.isna()
    df = df.loc[mask_keep].copy()
    df['stt'] = stt_numeric.loc[df.index]
    
    # Tạo cột note để lưu giá trị không hợp lệ
    df['note'] = ''
    
    # Lưu giá trị không hợp lệ vào cột note trước khi chuyển đổi
    for col in ['can_nang', 'chieu_cao', 'thang_tuoi']:
        # Tìm các giá trị không phải số
        invalid_mask = pd.to_numeric(df[col], errors='coerce').isna() & df[col].notna()
        # Copy giá trị không hợp lệ vào note
        df.loc[invalid_mask, 'note'] = df.loc[invalid_mask, 'note'] + df.loc[invalid_mask, col].astype(str) + '; '
    
    # Chuyển các cột số về đúng kiểu dữ liệu (giá trị không hợp lệ thành NaN)
    df['stt'] = pd.to_numeric(df['stt'], errors='coerce')
    df['thang_tuoi'] = pd.to_numeric(df['thang_tuoi'], errors='coerce')
    df['can_nang'] = pd.to_numeric(df['can_nang'], errors='coerce')
    df['chieu_cao'] = pd.to_numeric(df['chieu_cao'], errors='coerce')
    
    # Xóa dấu ; thừa ở cuối note
    df['note'] = df['note'].str.rstrip('; ')
    
    # Thêm cột giới tính dựa trên cột nam/nu
    df['gioi_tinh'] = df.apply(
        lambda row: 'trai' if pd.notna(row['nam']) and row['nam'] == 'x' else 'gai', axis=1
    )
    
    # Reset index
    df = df.reset_index(drop=True)
    
    return df


def export_to_excel(
    df_result: pd.DataFrame,
    input_file: Optional[str] = None,
    output_file: Optional[str] = None,
    data_start_row: int = 7,
    sheet_name: Optional[str] = None,
    under5_table: Optional[pd.DataFrame] = None,
    under5_sheet_name: str = 'Thong ke <5T'
) -> None:
    """
    Copy file Excel gốc và ghi kết quả vào, giữ nguyên định dạng.
    
    Args:
        df_result: DataFrame chứa kết quả cần ghi
        input_file: File Excel gốc làm template
        output_file: File Excel đầu ra (None = ghi đè file gốc)
        data_start_row: Hàng bắt đầu ghi data (mặc định là 8)
        sheet_name: Tên sheet (None = sheet đầu tiên)
    """
    # Nếu không có output_file, ghi đè file gốc
    if output_file is None:
        output_file = input_file
    
    # Chuẩn hóa đường dẫn
    input_file = os.path.normpath(input_file)
    output_file = os.path.normpath(output_file)
    
    # Copy file gốc sang file mới (giữ nguyên định dạng)
    if input_file != output_file:
        shutil.copy2(input_file, output_file)
    
    # Mở file bằng openpyxl
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active if sheet_name is None else wb[sheet_name]
    
    # Mapping cột DataFrame -> cột Excel (format mới 2025)
    column_mapping = {
        'stt': 'A',
        'ho_ten': 'B',
        'nam': 'C',
        'nu': 'D',
        'ngay_sinh': 'E',
        'thang_tuoi': 'F',
        'ho_ten_me': 'G',
        'can_nang': 'H',
        'execute_cn_tuoi': 'I',
        'chieu_cao': 'J',
        'execute_cc_tuoi': 'K',
        'execute_cn_cc': 'L',
        'note': 'M',
    }
    
    from process_helper import _strip_accents

    # Ghi từng dòng data
    for idx, row in df_result.iterrows():
        excel_row = int(data_start_row) + int(idx)

        for col_name, excel_col in column_mapping.items():
            if col_name in row.index:
                value = row[col_name]
                target = ws[f"{excel_col}{excel_row}"]
                # Xử lý giá trị NaN
                if pd.isna(value):
                    # Giữ 'Vắng' nếu ô template đang chứa -> không ghi đè rỗng
                    if isinstance(target.value, str) and 'vang' in _strip_accents(target.value):
                        continue
                    cell_value = None
                elif col_name in ['stt', 'thang_tuoi'] and pd.notna(value):
                    cell_value = int(value)
                else:
                    cell_value = value

                target.value = cell_value
    
    # Lưu file
    # Thêm sheet thống kê <5T nếu có
    if under5_table is not None:
        if under5_sheet_name in wb.sheetnames:
            del wb[under5_sheet_name]
        ws_stat = wb.create_sheet(under5_sheet_name)
        # Header
        for col_idx, col_name in enumerate(under5_table.columns, start=1):
            ws_stat.cell(row=1, column=col_idx, value=col_name)
        # Data
        for row_idx, row in enumerate(under5_table.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws_stat.cell(row=row_idx, column=col_idx, value=value)

    wb.save(output_file)
    wb.close()


def build_under5_statistics_table(summary: dict) -> pd.DataFrame:
    """Tạo bảng thống kê trẻ <5 tuổi (Tổng/Nam/Nữ) theo cấu trúc yêu cầu."""

    def get(section: str, gender: str | None) -> int:
        # Lấy giá trị theo giới tính nếu có, ngược lại lấy tổng
        if gender:
            return summary.get(section, {}).get(gender, 0)
        return summary.get(section, {}).get('tong', 0)

    def rate(numerator: int, denominator: int) -> float:
        return round(numerator / denominator * 100, 2) if denominator else 0.0

    def make_row(stt, label: str, gender: str | None):
        total = get('tong_so_tre', gender)
        can_do = get('tre_duoc_can', gender)
        sdd_cn_tuoi = get('sdd_cn_tuoi', gender)
        sdd_cc_tuoi = get('sdd_cc_tuoi', gender)
        sdd_cn_cc = get('sdd_cn_cc', gender)
        thua_can_bp = get('thua_can_beo_phi', gender)

        return {
            'STT': stt,
            'Chỉ số': label,
            'Số trẻ dưới 5 tuổi': total,
            'Số trẻ dưới 5 tuổi được cân, đo': can_do,
            'Số TE <5T bị SDD thể nhẹ cân (CN/T)': sdd_cn_tuoi,
            'Số TE <5T bị SDD thể thấp còi (CC/T)': sdd_cc_tuoi,
            'Số TE <5 tuổi bị SDD thể gầy còm (CN/CC)': sdd_cn_cc,
            'Số TE <5 tuổi bị Thừa cân, béo phì': thua_can_bp,
            'Tỷ lệ Cân nặng/tuổi (thể nhẹ cân)': rate(sdd_cn_tuoi, total),
            'Tỷ lệ Chiều cao/tuổi (thể thấp còi)': rate(sdd_cc_tuoi, total),
            'Tỷ lệ Cân nặng/chiều cao (thể gầy còm)': rate(sdd_cn_cc, total),
            'Tỷ lệ Thừa cân, Béo phì': rate(thua_can_bp, total),
        }

    rows = [
        make_row('', 'Tổng số', None),
        make_row(1, 'Nam', 'trai'),
        make_row(2, 'Nữ', 'gai'),
    ]

    return pd.DataFrame(rows)


def export_under5_statistics_sheet(summary: dict, excel_path: str, sheet_name: str = 'Thong ke <5T') -> None:
    """Ghi thêm sheet thống kê <5 tuổi (Tổng/Nam/Nữ) vào file Excel."""
    df_table = build_under5_statistics_table(summary)

    # Load workbook (tạo mới sheet, nếu đã tồn tại thì xóa để ghi lại)
    wb = openpyxl.load_workbook(excel_path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # Ghi header
    for col_idx, col_name in enumerate(df_table.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Ghi data
    for row_idx, row in enumerate(df_table.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(excel_path)
    wb.close()
        
