from typing import Optional
import pandas as pd
import numpy as np
import unicodedata
from datetime import datetime
from openpyxl import load_workbook


def _strip_accents(s) -> str:
    """Bỏ dấu tiếng Việt + lowercase để so khớp không phân biệt dấu/hoa thường."""
    s = unicodedata.normalize('NFD', str(s))
    return ''.join(c for c in s if unicodedata.category(c) != 'Mn').lower()


def _absent_mask(df: pd.DataFrame) -> np.ndarray:
    """Trẻ VẮNG MẶT: có chữ 'vắng' (mọi biến thể dấu/hoa thường: vang, vắng, Vắng, VẮNG...)
    ở BẤT KỲ cột nào được load lên. Trả mảng bool theo thứ tự dòng của df."""
    absent = pd.Series(False, index=df.index)
    for c in df.columns:
        col = df[c]
        if col.dtype == object:  # chỉ xét cột chứa text
            absent = absent | col.map(
                lambda v: isinstance(v, str) and 'vang' in _strip_accents(v)
            )
    return absent.to_numpy()


def calculate_month_age(birth_date: datetime, target_date: datetime) -> Optional[int]:
    if pd.isna(birth_date):
        return None
    
    birth = pd.to_datetime(birth_date)
    target = pd.to_datetime(target_date)
    
    # Tính số tháng (chỉ quan tâm năm và tháng)
    months = (target.year - birth.year) * 12 + (target.month - birth.month)
    
    return months


def apply_month_age(df: pd.DataFrame, target_date, birth_col: str = 'ngay_sinh', result_col: str = 'thang_tuoi') -> pd.DataFrame:
    df = df.copy()
    df[result_col] = df[birth_col].apply(lambda x: calculate_month_age(x, target_date))
    return df


def round_height_to_half(df: pd.DataFrame, height_col: str = 'chieu_cao') -> pd.DataFrame:
    """
    La  m tro  n chi eu cao ve  bu o c 0.5 cm de d o ng bo  du  lieu.
    NaN du o  c giu la i.
    """
    if height_col not in df.columns:
        return df.copy()
    
    df_rounded = df.copy()
    df_rounded[height_col] = df_rounded[height_col].apply(
        lambda x: np.nan if pd.isna(x) else np.round(x * 2) / 2
    )
    return df_rounded


def execute_weight_by_age(df_children: pd.DataFrame, df_weight_by_age: pd.DataFrame) -> pd.DataFrame:
    df = pd.merge(df_children, df_weight_by_age, how='left', left_on=['thang_tuoi', 'gioi_tinh'], right_on=['thang_tuoi', 'gioi_tinh'])
    
    conditions = [
        df['can_nang'] < df['minus_3sd'],
        df['can_nang'] < df['minus_2sd'],
        df['can_nang'] <= df['plus_2sd'],
        df['can_nang'] <= df['plus_3sd'],
        df['can_nang'] > df['plus_3sd'],
    ]
    choices = ['-3 SD', '-2 SD', 'BT', '+2 SD', '+3 SD']
    
    df['execute_cn_tuoi'] = np.select(conditions, choices, default='')
    
    df = df.drop(columns=df_weight_by_age.columns.difference(['thang_tuoi', 'gioi_tinh']))
    
    return df


def execute_height_by_age(df_children: pd.DataFrame, df_height_by_age: pd.DataFrame) -> pd.DataFrame:
    df = pd.merge(df_children, df_height_by_age, how='left', left_on=['thang_tuoi', 'gioi_tinh'], right_on=['thang_tuoi', 'gioi_tinh'])
    
    conditions = [
        df['chieu_cao'] < df['minus_3sd'],
        df['chieu_cao'] < df['minus_2sd'],
        df['chieu_cao'] <= df['plus_2sd'],
        df['chieu_cao'] <= df['plus_3sd'],
        df['chieu_cao'] > df['plus_3sd'],
    ]
    choices = ['-3 SD', '-2 SD', 'BT', '+2 SD', '+3 SD']
    
    df['execute_cc_tuoi'] = np.select(conditions, choices, default='')
    
    # Bỏ các cột tham chiếu
    df = df.drop(columns=df_height_by_age.columns.difference(['thang_tuoi', 'gioi_tinh']))
    
    return df


def execute_weight_by_height(df_children: pd.DataFrame, df_weight_by_height_0_2: pd.DataFrame, df_weight_by_height_2_5: pd.DataFrame) -> pd.DataFrame:
    # Tách những hàng có chieu_cao và không có
    df_with_height = df_children[df_children['chieu_cao'].notna()].copy()
    df_without_height = df_children[df_children['chieu_cao'].isna()].copy()
    
    # Sắp xếp theo chieu_cao để dùng merge_asof
    df_with_height = df_with_height.sort_values('chieu_cao')
    
    # Tách trẻ 0-24 tháng và trẻ > 24 tháng
    df_0_24 = df_with_height[df_with_height['thang_tuoi'] <= 24]
    df_25_60 = df_with_height[df_with_height['thang_tuoi'] > 24]
    
    results = []
    
    # Merge_asof cho từng giới tính (0-24 tháng)
    for gioi_tinh in ['trai', 'gai']:
        df_child = df_0_24[df_0_24['gioi_tinh'] == gioi_tinh]
        df_ref = df_weight_by_height_0_2[df_weight_by_height_0_2['gioi_tinh'] == gioi_tinh].sort_values('chieu_cao')
        if not df_child.empty:
            merged = pd.merge_asof(df_child, df_ref, on='chieu_cao', direction='backward', suffixes=('', '_ref'))
            merged = merged.drop(columns=['gioi_tinh_ref'], errors='ignore')
            results.append(merged)
    
    # Merge_asof cho từng giới tính (25-60 tháng)
    for gioi_tinh in ['trai', 'gai']:
        df_child = df_25_60[df_25_60['gioi_tinh'] == gioi_tinh]
        df_ref = df_weight_by_height_2_5[df_weight_by_height_2_5['gioi_tinh'] == gioi_tinh].sort_values('chieu_cao')
        if not df_child.empty:
            merged = pd.merge_asof(df_child, df_ref, on='chieu_cao', direction='backward', suffixes=('', '_ref'))
            merged = merged.drop(columns=['gioi_tinh_ref'], errors='ignore')
            results.append(merged)
    
    # Gộp lại với những hàng không có chiều cao
    results.append(df_without_height)
    df = pd.concat(results, ignore_index=True)
    
    conditions = [
        df['can_nang'] < df['minus_3sd'],
        df['can_nang'] < df['minus_2sd'],
        df['can_nang'] <= df['plus_2sd'],
        df['can_nang'] <= df['plus_3sd'],
        df['can_nang'] > df['plus_3sd'],
    ]
    choices = ['-3 SD', '-2 SD', 'BT', '+2 SD', '+3 SD']
    
    df['execute_cn_cc'] = np.select(conditions, choices, default='')
    
    # Bỏ các cột tham chiếu
    df = df.drop(columns=df_weight_by_height_0_2.columns.difference(['chieu_cao', 'gioi_tinh']))
    
    # Sắp xếp lại theo stt và reset index
    df = df.sort_values('stt').reset_index(drop=True)
    
    return df


def fill_missing_height_normal(
    df_children: pd.DataFrame,
    df_height_by_age: pd.DataFrame,
    offset_range: tuple = (0.0, 2.0)
) -> pd.DataFrame:
    """
    Điền chiều cao cho trẻ CHƯA có số đo VÀ CHƯA có trạng thái CC/tuổi.
    Giá trị điền nằm sát đầu thấp của khoảng bình thường (BT):
    minus_2sd + random nhỏ, kẹp <= plus_2sd -> luôn rơi vào BT.
    Chỉ điền số đo, KHÔNG ghi trạng thái (để bước Chấm SDD tự tính nếu cần).
    """
    if 'chieu_cao' not in df_children.columns:
        return df_children
    df = df_children.copy()

    if 'execute_cc_tuoi' in df.columns:
        no_status = df['execute_cc_tuoi'].fillna('').astype(str).str.strip() == ''
    else:
        no_status = pd.Series(True, index=df.index)
    mask = (df['chieu_cao'].isna() & no_status).to_numpy() & ~_absent_mask(df)
    if not mask.any():
        return df_children

    # Tra chuẩn WHO theo (thang_tuoi, gioi_tinh) — reindex giữ ĐÚNG thứ tự & độ dài của df
    ref = df_height_by_age.drop_duplicates(['thang_tuoi', 'gioi_tinh']).set_index(['thang_tuoi', 'gioi_tinh'])
    keys = pd.MultiIndex.from_arrays([df['thang_tuoi'], df['gioi_tinh']])
    lo = ref['minus_2sd'].reindex(keys).to_numpy(dtype='float64')
    hi = ref['plus_2sd'].reindex(keys).to_numpy(dtype='float64')

    rnd = np.random.uniform(offset_range[0], offset_range[1], len(df))
    new_vals = np.round(np.minimum(lo + rnd, hi), 1)  # kẹp <= plus_2sd

    fill = mask & ~np.isnan(lo)  # chỉ điền khi có chuẩn WHO tương ứng
    arr = df['chieu_cao'].to_numpy(dtype='float64')
    arr[fill] = new_vals[fill]
    df['chieu_cao'] = arr
    return df


def fill_missing_weight_normal(
    df_children: pd.DataFrame,
    df_weight_by_age: pd.DataFrame,
    offset_range: tuple = (0.0, 1.0)
) -> pd.DataFrame:
    """
    Điền cân nặng cho trẻ CHƯA có số đo VÀ CHƯA có trạng thái CN/tuổi.
    Dựa theo chuẩn CN/tuổi (không cần chiều cao). Giá trị điền sát đầu thấp
    khoảng BT: minus_2sd + random nhỏ, kẹp <= plus_2sd. Chỉ điền số, không ghi trạng thái.
    """
    if 'can_nang' not in df_children.columns:
        return df_children
    df = df_children.copy()

    if 'execute_cn_tuoi' in df.columns:
        no_status = df['execute_cn_tuoi'].fillna('').astype(str).str.strip() == ''
    else:
        no_status = pd.Series(True, index=df.index)
    mask = (df['can_nang'].isna() & no_status).to_numpy() & ~_absent_mask(df)
    if not mask.any():
        return df_children

    # Tra chuẩn WHO theo (thang_tuoi, gioi_tinh) — reindex giữ ĐÚNG thứ tự & độ dài của df
    ref = df_weight_by_age.drop_duplicates(['thang_tuoi', 'gioi_tinh']).set_index(['thang_tuoi', 'gioi_tinh'])
    keys = pd.MultiIndex.from_arrays([df['thang_tuoi'], df['gioi_tinh']])
    lo = ref['minus_2sd'].reindex(keys).to_numpy(dtype='float64')
    hi = ref['plus_2sd'].reindex(keys).to_numpy(dtype='float64')

    rnd = np.random.uniform(offset_range[0], offset_range[1], len(df))
    new_vals = np.round(np.minimum(lo + rnd, hi), 1)

    fill = mask & ~np.isnan(lo)
    arr = df['can_nang'].to_numpy(dtype='float64')
    arr[fill] = new_vals[fill]
    df['can_nang'] = arr
    return df


def _random_value_for_status(status, m3, m2, p2, p3):
    """Sinh 1 giá trị nằm TRONG band của trạng thái (theo đúng quy ước hàm chấm:
    <m3=-3SD, [m3,m2)=-2SD, [m2,p2]=BT, (p2,p3]=+2SD, >p3=+3SD).
    Trả None nếu thiếu ngưỡng chuẩn."""
    if status == '-3 SD':
        if np.isnan(m3):
            return None
        return round(max(m3 - np.random.uniform(0.3, 1.0), 0.1), 1)
    if status == '-2 SD':
        if np.isnan(m3) or np.isnan(m2) or m2 <= m3:
            return None
        v = round(m3 + (m2 - m3) * np.random.uniform(0.2, 0.8), 1)
        return min(v, round(m2 - 0.1, 1))  # đảm bảo < m2
    if status == 'BT':
        if np.isnan(m2) or np.isnan(p2) or p2 <= m2:
            return None
        return round(m2 + (p2 - m2) * np.random.uniform(0.05, 0.4), 1)  # gần đầu thấp BT
    if status == '+2 SD':
        if np.isnan(p2) or np.isnan(p3) or p3 <= p2:
            return None
        return max(round(p2 + (p3 - p2) * np.random.uniform(0.2, 0.8), 1), round(p2 + 0.1, 1))
    if status == '+3 SD':
        if np.isnan(p3):
            return None
        return round(p3 + np.random.uniform(0.3, 1.0), 1)
    return None


def _fill_by_status(df_children, df_standard, value_col, status_col):
    """Điền số đo cho trẻ CHƯA có số (NaN) NHƯNG ĐÃ CÓ trạng thái ghi sẵn.
    Giá trị điền nằm trong band của trạng thái. KHÔNG đụng vào số đo đã có."""
    if value_col not in df_children.columns or status_col not in df_children.columns:
        return df_children
    df = df_children.copy()
    valid = ['-3 SD', '-2 SD', 'BT', '+2 SD', '+3 SD']
    status = df[status_col].fillna('').astype(str).str.strip()
    # Bỏ qua trẻ VẮNG MẶT (có chữ 'vắng' ở bất kỳ cột nào)
    mask = (df[value_col].isna() & status.isin(valid)).to_numpy() & ~_absent_mask(df)
    if not mask.any():
        return df_children

    ref = df_standard.drop_duplicates(['thang_tuoi', 'gioi_tinh']).set_index(['thang_tuoi', 'gioi_tinh'])
    keys = pd.MultiIndex.from_arrays([df['thang_tuoi'], df['gioi_tinh']])
    m3 = ref['minus_3sd'].reindex(keys).to_numpy('float64')
    m2 = ref['minus_2sd'].reindex(keys).to_numpy('float64')
    p2 = ref['plus_2sd'].reindex(keys).to_numpy('float64')
    p3 = ref['plus_3sd'].reindex(keys).to_numpy('float64')
    st = status.to_numpy()

    arr = df[value_col].to_numpy(dtype='float64')
    for i in np.flatnonzero(mask):
        v = _random_value_for_status(st[i], m3[i], m2[i], p2[i], p3[i])
        if v is not None:
            arr[i] = v
    df[value_col] = arr
    return df


def fill_height_by_status(df_children: pd.DataFrame, df_height_by_age: pd.DataFrame) -> pd.DataFrame:
    """Điền chiều cao cho trẻ CHƯA có số đo nhưng ĐÃ CÓ trạng thái CC/tuổi (execute_cc_tuoi).
    Giá trị khớp band trạng thái. Giữ nguyên số đo gốc."""
    return _fill_by_status(df_children, df_height_by_age, 'chieu_cao', 'execute_cc_tuoi')


def fill_weight_by_status(df_children: pd.DataFrame, df_weight_by_age: pd.DataFrame) -> pd.DataFrame:
    """Điền cân nặng cho trẻ CHƯA có số đo nhưng ĐÃ CÓ trạng thái CN/tuổi (execute_cn_tuoi).
    Theo chuẩn CN/tuổi, khớp band trạng thái. Giữ nguyên số đo gốc."""
    return _fill_by_status(df_children, df_weight_by_age, 'can_nang', 'execute_cn_tuoi')


# ===== Helper cho fill_cn_cc =====
_SD_LABELS = ['-3 SD', '-2 SD', 'BT', '+2 SD', '+3 SD']


def _classify_sd(x, m3, m2, p2, p3):
    """Phân loại 1 giá trị theo đúng quy ước hàm chấm."""
    if pd.isna(x):
        return ''
    if x < m3:
        return '-3 SD'
    if x < m2:
        return '-2 SD'
    if x <= p2:
        return 'BT'
    if x <= p3:
        return '+2 SD'
    return '+3 SD'


def _band_sd(status, m3, m2, p2, p3):
    """(lo, hi) của band ứng với trạng thái (dùng -inf/inf cho đầu mở)."""
    if status == '-3 SD':
        return (float('-inf'), m3)
    if status == '-2 SD':
        return (m3, m2)
    if status == '+2 SD':
        return (p2, p3)
    if status == '+3 SD':
        return (p3, float('inf'))
    return (m2, p2)  # BT (mặc định)


def _pick_in_band(lo, hi):
    """Chọn 1 giá trị nằm trong band, thiên về đầu thấp; None nếu band vô hạn 2 đầu."""
    if lo == float('-inf') and hi == float('inf'):
        return None
    if lo == float('-inf'):
        return round(hi - np.random.uniform(0.3, 1.0), 1)
    if hi == float('inf'):
        return round(lo + np.random.uniform(0.3, 1.0), 1)
    if hi <= lo:
        return round(lo, 1)
    return round(lo + (hi - lo) * np.random.uniform(0.05, 0.4), 1)


def _intersect_band(b1, b2):
    """Giao 2 band; None nếu rỗng."""
    lo = max(b1[0], b2[0])
    hi = min(b1[1], b2[1])
    return (lo, hi) if lo <= hi else None


def _age_thresholds(std, tt, g):
    sub = std[(std['thang_tuoi'] == tt) & (std['gioi_tinh'] == g)]
    if sub.empty:
        return None
    r = sub.iloc[0]
    return (r['minus_3sd'], r['minus_2sd'], r['plus_2sd'], r['plus_3sd'])


def _cncc_thresholds(height, tt, g, t02, t25):
    """Ngưỡng CN/CC tại 1 chiều cao (tra bảng theo nhóm tuổi + giới, lùi về mốc gần nhất)."""
    if pd.isna(height) or pd.isna(tt):
        return None
    tbl = t02 if tt <= 24 else t25
    sub = tbl[tbl['gioi_tinh'] == g].sort_values('chieu_cao')
    if sub.empty:
        return None
    le = sub[sub['chieu_cao'] <= height]
    r = le.iloc[-1] if not le.empty else sub.iloc[0]
    return (r['minus_3sd'], r['minus_2sd'], r['plus_2sd'], r['plus_3sd'])


def _valid_or_bt(s):
    """Chỉ tiêu trống/không hợp lệ -> coi là BT."""
    s = '' if s is None else str(s).strip()
    return s if s in _SD_LABELS else 'BT'


def fill_cn_cc(
    df_children: pd.DataFrame,
    df_weight_by_height_0_2: pd.DataFrame,
    df_weight_by_height_2_5: pd.DataFrame,
    df_height_by_age: pd.DataFrame,
    df_weight_by_age: pd.DataFrame,
) -> pd.DataFrame:
    """
    Điền chiều cao và/hoặc cân nặng để đạt CHỈ TIÊU CN/CC (execute_cn_cc),
    cố gắng giữ chỉ tiêu CN/tuổi (execute_cn_tuoi) & CC/tuổi (execute_cc_tuoi),
    nhưng UỐN trong band để điền được CN/CC (ưu tiên CN/CC).

    Quy tắc:
    - Chỉ điền ô CÒN TRỐNG (giữ nguyên số đo đã có).
    - CHỈ điền khi: ô CN/CC CÓ trạng thái hợp lệ VÀ CẢ HAI số đo (CN, CC) đều trống.
      (Nếu CN/CC trống, hoặc đã có sẵn 1 trong 2 số đo -> BỎ QUA, giữ nguyên.)
    - Điền cả chiều cao lẫn cân nặng để đạt chỉ tiêu CN/CC.
    - Chỉ tiêu CN/tuổi & CC/tuổi để trống -> coi là BT (ràng buộc mềm).
    - Trẻ 'vắng' (mọi cột) -> bỏ qua.
    """
    needed = ['chieu_cao', 'can_nang', 'thang_tuoi', 'gioi_tinh']
    if any(c not in df_children.columns for c in needed):
        return df_children

    df = df_children.copy()
    absent = _absent_mask(df)
    cc = df['chieu_cao'].to_numpy(dtype='float64')
    cn = df['can_nang'].to_numpy(dtype='float64')
    tt_arr = df['thang_tuoi'].to_numpy(dtype='float64')
    g_arr = df['gioi_tinh'].to_numpy()

    def col_st(name):
        if name in df.columns:
            return df[name].fillna('').astype(str).str.strip().to_numpy()
        return np.array([''] * len(df), dtype=object)

    cc_st = col_st('execute_cc_tuoi')
    cn_st = col_st('execute_cn_tuoi')
    cncc_st = col_st('execute_cn_cc')

    t02, t25 = df_weight_by_height_0_2, df_weight_by_height_2_5

    for i in range(len(df)):
        if absent[i]:
            continue
        has_cc = not np.isnan(cc[i])
        has_cn = not np.isnan(cn[i])
        # CHỈ điền khi CẢ HAI số đo (cân nặng VÀ chiều cao) đều TRỐNG
        if has_cc or has_cn:
            continue
        tt = tt_arr[i]
        g = g_arr[i]
        if np.isnan(tt):
            continue
        # Chỉ điền khi ô CN/CC CÓ trạng thái hợp lệ; CN/CC trống -> BỎ QUA
        cncc_t = str(cncc_st[i]).strip()
        if cncc_t not in _SD_LABELS:
            continue
        cc_t = _valid_or_bt(cc_st[i])
        cn_t = _valid_or_bt(cn_st[i])

        # Điền chiều cao theo band CC/tuổi, rồi cân nặng theo band CN/CC ∩ CN/tuổi
        hthr = _age_thresholds(df_height_by_age, tt, g)
        h = _pick_in_band(*_band_sd(cc_t, *hthr)) if hthr else None
        if h is None:
            continue
        cc[i] = h
        thr = _cncc_thresholds(h, tt, g, t02, t25)
        if thr is None:
            continue
        band = _band_sd(cncc_t, *thr)
        wthr = _age_thresholds(df_weight_by_age, tt, g)
        if wthr:
            inter = _intersect_band(band, _band_sd(cn_t, *wthr))
            if inter:
                band = inter
        v = _pick_in_band(*band)
        if v is not None:
            cn[i] = v

    df['chieu_cao'] = cc
    df['can_nang'] = cn
    return df


def summary_statistics(df: pd.DataFrame, max_months: Optional[int] = None) -> dict:
    """
    Tổng kết dinh dưỡng
    Args:
        df: DataFrame chứa dữ liệu trẻ em
        max_months: Giới hạn tháng tuổi (None = tất cả, 24 = dưới 2 tuổi, 60 = dưới 5 tuổi)
    """
    # Lọc theo tháng tuổi nếu có
    if max_months:
        df_filtered = df[df['thang_tuoi'] <= max_months].copy()
    else:
        df_filtered = df.copy()
    
    # === TỔNG SỐ TRẺ ===
    total = len(df_filtered)
    total_trai = len(df_filtered[df_filtered['gioi_tinh'] == 'trai'])
    total_gai = len(df_filtered[df_filtered['gioi_tinh'] == 'gai'])
    
    # === TRẺ ĐƯỢC CÂN ===
    df_can = df_filtered[df_filtered['can_nang'].notna()]
    can_trai = len(df_can[df_can['gioi_tinh'] == 'trai'])
    can_gai = len(df_can[df_can['gioi_tinh'] == 'gai'])
    ty_le_can = len(df_can) / total * 100 if total > 0 else 0
    
    # === TRẺ ĐƯỢC ĐO ===
    df_do = df_filtered[df_filtered['chieu_cao'].notna()]
    do_trai = len(df_do[df_do['gioi_tinh'] == 'trai'])
    do_gai = len(df_do[df_do['gioi_tinh'] == 'gai'])
    ty_le_do = len(df_do) / total * 100 if total > 0 else 0
    
    # === SDD CÂN NẶNG/TUỔI ===
    # SDD: -2SD và -3SD
    sdd_cn_tuoi = df_can[df_can['execute_cn_tuoi'].isin(['-2 SD', '-3 SD'])]
    sdd_cn_tuoi_trai = len(sdd_cn_tuoi[sdd_cn_tuoi['gioi_tinh'] == 'trai'])
    sdd_cn_tuoi_gai = len(sdd_cn_tuoi[sdd_cn_tuoi['gioi_tinh'] == 'gai'])
    sdd_cn_tuoi_2sd = len(df_can[df_can['execute_cn_tuoi'] == '-2 SD'])
    sdd_cn_tuoi_3sd = len(df_can[df_can['execute_cn_tuoi'] == '-3 SD'])
    # Thừa cân: +2SD, +3SD
    thua_cn_tuoi = df_can[df_can['execute_cn_tuoi'].isin(['+2 SD', '+3 SD'])]
    thua_cn_tuoi_trai = len(thua_cn_tuoi[thua_cn_tuoi['gioi_tinh'] == 'trai'])
    thua_cn_tuoi_gai = len(thua_cn_tuoi[thua_cn_tuoi['gioi_tinh'] == 'gai'])
    ty_le_sdd_cn_tuoi = len(sdd_cn_tuoi) / len(df_can) * 100 if len(df_can) > 0 else 0
    
    # === SDD CHIỀU CAO/TUỔI ===
    sdd_cc_tuoi = df_do[df_do['execute_cc_tuoi'].isin(['-2 SD', '-3 SD'])]
    sdd_cc_tuoi_trai = len(sdd_cc_tuoi[sdd_cc_tuoi['gioi_tinh'] == 'trai'])
    sdd_cc_tuoi_gai = len(sdd_cc_tuoi[sdd_cc_tuoi['gioi_tinh'] == 'gai'])
    sdd_cc_tuoi_2sd = len(df_do[df_do['execute_cc_tuoi'] == '-2 SD'])
    sdd_cc_tuoi_3sd = len(df_do[df_do['execute_cc_tuoi'] == '-3 SD'])
    ty_le_sdd_cc_tuoi = len(sdd_cc_tuoi) / len(df_do) * 100 if len(df_do) > 0 else 0
    
    # === SDD CÂN NẶNG/CHIỀU CAO ===
    # Mẫu số = số trẻ ĐƯỢC ĐÁNH GIÁ CN/CC (có phân loại hợp lệ), KHÔNG tính ô rỗng/NaN
    valid_cn_cc_labels = ['-3 SD', '-2 SD', 'BT', '+2 SD', '+3 SD']
    df_cn_cc = df_filtered[df_filtered['execute_cn_cc'].isin(valid_cn_cc_labels)]
    sdd_cn_cc = df_cn_cc[df_cn_cc['execute_cn_cc'].isin(['-2 SD', '-3 SD'])]
    sdd_cn_cc_trai = len(sdd_cn_cc[sdd_cn_cc['gioi_tinh'] == 'trai'])
    sdd_cn_cc_gai = len(sdd_cn_cc[sdd_cn_cc['gioi_tinh'] == 'gai'])
    sdd_cn_cc_2sd = len(df_cn_cc[df_cn_cc['execute_cn_cc'] == '-2 SD'])
    sdd_cn_cc_3sd = len(df_cn_cc[df_cn_cc['execute_cn_cc'] == '-3 SD'])
    so_danh_gia_cn_cc = len(df_cn_cc)
    ty_le_sdd_cn_cc = len(sdd_cn_cc) / so_danh_gia_cn_cc * 100 if so_danh_gia_cn_cc > 0 else 0
    
    # === THỪA CÂN VÀ BÉO PHÌ (CN/Tuổi) ===
    # Tổng thừa cân + béo phì (+2SD và +3SD)
    thua_can_beo_phi = df_can[df_can['execute_cn_tuoi'].isin(['+2 SD', '+3 SD'])]
    thua_can_beo_phi_tong = len(thua_can_beo_phi)
    thua_can_beo_phi_trai = len(thua_can_beo_phi[thua_can_beo_phi['gioi_tinh'] == 'trai'])
    thua_can_beo_phi_gai = len(thua_can_beo_phi[thua_can_beo_phi['gioi_tinh'] == 'gai'])
    
    # Đếm riêng mức +2SD và +3SD
    thua_can_beo_phi_2sd = len(df_can[df_can['execute_cn_tuoi'] == '+2 SD'])
    thua_can_beo_phi_3sd = len(df_can[df_can['execute_cn_tuoi'] == '+3 SD'])
    
    # Thừa cân: +2SD (giữ lại để tương thích)
    thua_can_cn_tuoi = df_can[df_can['execute_cn_tuoi'] == '+2 SD']
    thua_can_cn_tuoi_tong = len(thua_can_cn_tuoi)
    thua_can_cn_tuoi_trai = len(thua_can_cn_tuoi[thua_can_cn_tuoi['gioi_tinh'] == 'trai'])
    thua_can_cn_tuoi_gai = len(thua_can_cn_tuoi[thua_can_cn_tuoi['gioi_tinh'] == 'gai'])
    
    # Béo phì: +3SD (giữ lại để tương thích)
    beo_phi_cn_tuoi = df_can[df_can['execute_cn_tuoi'] == '+3 SD']
    beo_phi_cn_tuoi_tong = len(beo_phi_cn_tuoi)
    beo_phi_cn_tuoi_trai = len(beo_phi_cn_tuoi[beo_phi_cn_tuoi['gioi_tinh'] == 'trai'])
    beo_phi_cn_tuoi_gai = len(beo_phi_cn_tuoi[beo_phi_cn_tuoi['gioi_tinh'] == 'gai'])
    
    result = {
        'tong_so_tre': {'tong': total, 'trai': total_trai, 'gai': total_gai},
        'tre_duoc_can': {
            'tong': len(df_can), 'trai': can_trai, 'gai': can_gai,
            'ty_le': round(ty_le_can, 2)
        },
        'tre_duoc_do': {
            'tong': len(df_do), 'trai': do_trai, 'gai': do_gai,
            'ty_le': round(ty_le_do, 2)
        },
        'sdd_cn_tuoi': {
            'tong': len(sdd_cn_tuoi), 'trai': sdd_cn_tuoi_trai, 'gai': sdd_cn_tuoi_gai,
            'muc_2sd': sdd_cn_tuoi_2sd, 'muc_3sd': sdd_cn_tuoi_3sd,
            'ty_le': round(ty_le_sdd_cn_tuoi, 2)
        },
        'sdd_cc_tuoi': {
            'tong': len(sdd_cc_tuoi), 'trai': sdd_cc_tuoi_trai, 'gai': sdd_cc_tuoi_gai,
            'muc_2sd': sdd_cc_tuoi_2sd, 'muc_3sd': sdd_cc_tuoi_3sd,
            'ty_le': round(ty_le_sdd_cc_tuoi, 2)
        },
        'sdd_cn_cc': {
            'tong': len(sdd_cn_cc), 'trai': sdd_cn_cc_trai, 'gai': sdd_cn_cc_gai,
            'muc_2sd': sdd_cn_cc_2sd, 'muc_3sd': sdd_cn_cc_3sd,
            'danh_gia': so_danh_gia_cn_cc,  # mẫu số: số trẻ được đánh giá CN/CC
            'ty_le': round(ty_le_sdd_cn_cc, 2)
        },
        'thua_can_cn_tuoi': {
            'tong': thua_can_cn_tuoi_tong, 'trai': thua_can_cn_tuoi_trai, 'gai': thua_can_cn_tuoi_gai
        },
        'beo_phi_cn_tuoi': {
            'tong': beo_phi_cn_tuoi_tong, 'trai': beo_phi_cn_tuoi_trai, 'gai': beo_phi_cn_tuoi_gai
        },
        'thua_can_beo_phi': {
            'tong': thua_can_beo_phi_tong, 'trai': thua_can_beo_phi_trai, 'gai': thua_can_beo_phi_gai,
            'muc_2sd': thua_can_beo_phi_2sd, 'muc_3sd': thua_can_beo_phi_3sd
        }
    }
    
    return result


def print_summary(summary: dict, title: str = "TRẺ DƯỚI 2 TUỔI"):
    """In báo cáo tổng kết"""
    print(f"\n{'='*60}")
    print(f"TỔNG KẾT DINH DƯỠNG - {title}")
    print(f"{'='*60}")
    
    s = summary
    print(f"\n1. TỔNG SỐ TRẺ: {s['tong_so_tre']['tong']} (Trai: {s['tong_so_tre']['trai']}, Gái: {s['tong_so_tre']['gai']})")
    
    print(f"\n2. TRẺ ĐƯỢC CÂN: {s['tre_duoc_can']['tong']} (Trai: {s['tre_duoc_can']['trai']}, Gái: {s['tre_duoc_can']['gai']})")
    print(f"   Tỷ lệ: {s['tre_duoc_can']['ty_le']}%")
    
    print(f"\n3. TRẺ ĐƯỢC ĐO: {s['tre_duoc_do']['tong']} (Trai: {s['tre_duoc_do']['trai']}, Gái: {s['tre_duoc_do']['gai']})")
    print(f"   Tỷ lệ: {s['tre_duoc_do']['ty_le']}%")
    
    print(f"\n4. SDD CÂN NẶNG/TUỔI: {s['sdd_cn_tuoi']['tong']} (Trai: {s['sdd_cn_tuoi']['trai']}, Gái: {s['sdd_cn_tuoi']['gai']})")
    print(f"   - Mức -2SD: {s['sdd_cn_tuoi']['muc_2sd']}, Mức -3SD: {s['sdd_cn_tuoi']['muc_3sd']}")
    print(f"   Tỷ lệ SDD: {s['sdd_cn_tuoi']['ty_le']}%")
    
    print(f"\n5. SDD CHIỀU CAO/TUỔI: {s['sdd_cc_tuoi']['tong']} (Trai: {s['sdd_cc_tuoi']['trai']}, Gái: {s['sdd_cc_tuoi']['gai']})")
    print(f"   - Mức -2SD: {s['sdd_cc_tuoi']['muc_2sd']}, Mức -3SD: {s['sdd_cc_tuoi']['muc_3sd']}")
    print(f"   Tỷ lệ SDD: {s['sdd_cc_tuoi']['ty_le']}%")
    
    print(f"\n6. SDD CÂN NẶNG/CHIỀU CAO: {s['sdd_cn_cc']['tong']} (Trai: {s['sdd_cn_cc']['trai']}, Gái: {s['sdd_cn_cc']['gai']})")
    print(f"   - Mức -2SD: {s['sdd_cn_cc']['muc_2sd']}, Mức -3SD: {s['sdd_cn_cc']['muc_3sd']}")
    
    print(f"\n7. THỪA CÂN, BÉO PHÌ (CN/Tuổi +2SD, +3SD):")
    tc = s['thua_can_beo_phi']
    print(f"   Tổng: {tc['tong']} (Trai: {tc['trai']}, Gái: {tc['gai']})")
    
    print(f"\n{'='*60}\n")


def summary_table(summary: dict) -> pd.DataFrame:
    """Chuyển summary dict thành DataFrame dạng bảng phẳng để in/excel - theo đúng thứ tự bảng mẫu."""
    s = summary
    total = s['tong_so_tre']['tong'] or 1  # tránh chia 0
    len_can = s['tre_duoc_can']['tong'] or 1
    len_do = s['tre_duoc_do']['tong'] or 1
    
    # Tính tỷ lệ được cân đo (trên tổng số trẻ)
    ty_le_duoc_can_do = round(len_can / total * 100, 2)

    row = {
        # 1. TỔNG SỐ TRẺ
        'Tổng số trẻ': s['tong_so_tre']['tong'],
        'TS Trai': s['tong_so_tre']['trai'],
        'TS Gái': s['tong_so_tre']['gai'],
        
        # 2. SỐ TRẺ ĐƯỢC CÂN, ĐO
        'Số trẻ được cân, đo': len_can,
        'Cân đo Trai': s['tre_duoc_can']['trai'],
        'Cân đo Gái': s['tre_duoc_can']['gai'],
        'Tỷ lệ được cân, đo (%)': ty_le_duoc_can_do,
        
        # 3. CẦN XÁC ĐỊNH TRẺ DƯỚI 5 TUỔI CÓ THỪA CÂN HOẶC BÉO PHÌ
        'Thừa cân/béo phì': s['thua_can_beo_phi']['tong'],
        'Thừa cân/béo phì Trai': s['thua_can_beo_phi']['trai'],
        'Thừa cân/béo phì Gái': s['thua_can_beo_phi']['gai'],
        'Tỷ lệ thừa cân/béo phì (%)': round(s['thua_can_beo_phi']['tong'] / len_can * 100, 2),
        
        # 3.1 Trong đó: Thừa cân (+2SD)
        'Thừa cân (+2SD)': s['thua_can_beo_phi']['muc_2sd'],
        
        # 3.2 Trong đó: Béo phì (+3SD)
        'Béo phì (+3SD)': s['thua_can_beo_phi']['muc_3sd'],
        
        # 1. SỐ TRẺ BỊ SDD CN/TUỔI (THỂ NHẸ CÂN)
        'SDD CN/tuổi (nhẹ cân)': s['sdd_cn_tuoi']['tong'],
        'SDD CN/tuổi Trai': s['sdd_cn_tuoi']['trai'],
        'SDD CN/tuổi Gái': s['sdd_cn_tuoi']['gai'],
        'Tỷ lệ SDD CN/tuổi (%)': s['sdd_cn_tuoi']['ty_le'],
        
        # 1.1 Trong đó: SDD nhẹ cân (-2SD)
        'SDD nhẹ cân (-2SD)': s['sdd_cn_tuoi']['muc_2sd'],
        
        # 1.2 Trong đó: SDD nặng cân (-3SD) 
        'SDD nặng cân (-3SD)': s['sdd_cn_tuoi']['muc_3sd'],
        
        # 4. SỐ TRẺ BỊ SDD CC/TUỔI (THỂ THẤP CÒI)
        'SDD CC/tuổi (thấp còi)': s['sdd_cc_tuoi']['tong'],
        'SDD CC/tuổi Trai': s['sdd_cc_tuoi']['trai'],
        'SDD CC/tuổi Gái': s['sdd_cc_tuoi']['gai'],
        'Tỷ lệ SDD CC/tuổi (%)': s['sdd_cc_tuoi']['ty_le'],
        
        # 4.1 Trong đó: SDD thấp còi (-2SD)
        'SDD thấp còi (-2SD)': s['sdd_cc_tuoi']['muc_2sd'],
        
        # 4.2 Trong đó: SDD thấp còi nặng (-3SD)
        'SDD thấp còi nặng (-3SD)': s['sdd_cc_tuoi']['muc_3sd'],
        
        # 5. SỐ TRẺ BỊ SDD CN/CC (THỂ GẦY CÒM)
        'SDD CN/CC (gầy còm)': s['sdd_cn_cc']['tong'],
        'SDD CN/CC Trai': s['sdd_cn_cc']['trai'],
        'SDD CN/CC Gái': s['sdd_cn_cc']['gai'],
        'Tỷ lệ SDD CN/CC (%)': s['sdd_cn_cc']['ty_le'],
        
        # 5.1 Trong đó: SDD gầy còm (-2SD)
        'SDD gầy còm (-2SD)': s['sdd_cn_cc']['muc_2sd'],
        
        # 5.2 Trong đó: SDD gầy còm nặng (-3SD)
        'SDD gầy còm nặng (-3SD)': s['sdd_cn_cc']['muc_3sd'],
    }

    return pd.DataFrame([row])


def combine_summaries(summaries: list[dict]) -> Optional[dict]:
    """Gộp nhiều summary (đã cùng cấu trúc) thành một summary cộng gộp.

    Cộng tất cả các chỉ số đếm, sau đó tính lại tỷ lệ dựa trên tổng gộp, không trung bình cộng.
    """
    if not summaries:
        return None

    def add(a: dict, b: dict, keys: list[str]):
        for k in keys:
            a[k] = a.get(k, 0) + b.get(k, 0)

    agg = {
        'tong_so_tre': {'tong': 0, 'trai': 0, 'gai': 0},
        'tre_duoc_can': {'tong': 0, 'trai': 0, 'gai': 0},
        'tre_duoc_do': {'tong': 0, 'trai': 0, 'gai': 0},
        'sdd_cn_tuoi': {'tong': 0, 'trai': 0, 'gai': 0, 'muc_2sd': 0, 'muc_3sd': 0},
        'sdd_cc_tuoi': {'tong': 0, 'trai': 0, 'gai': 0, 'muc_2sd': 0, 'muc_3sd': 0},
        'sdd_cn_cc': {'tong': 0, 'trai': 0, 'gai': 0, 'muc_2sd': 0, 'muc_3sd': 0, 'danh_gia': 0},
        'thua_can_cn_tuoi': {'tong': 0, 'trai': 0, 'gai': 0},
        'beo_phi_cn_tuoi': {'tong': 0, 'trai': 0, 'gai': 0},
        'thua_can_beo_phi': {'tong': 0, 'trai': 0, 'gai': 0, 'muc_2sd': 0, 'muc_3sd': 0}
    }

    for s in summaries:
        add(agg['tong_so_tre'], s['tong_so_tre'], ['tong', 'trai', 'gai'])
        add(agg['tre_duoc_can'], s['tre_duoc_can'], ['tong', 'trai', 'gai'])
        add(agg['tre_duoc_do'], s['tre_duoc_do'], ['tong', 'trai', 'gai'])

        add(agg['sdd_cn_tuoi'], s['sdd_cn_tuoi'], ['tong', 'trai', 'gai', 'muc_2sd', 'muc_3sd'])
        add(agg['sdd_cc_tuoi'], s['sdd_cc_tuoi'], ['tong', 'trai', 'gai', 'muc_2sd', 'muc_3sd'])
        add(agg['sdd_cn_cc'], s['sdd_cn_cc'], ['tong', 'trai', 'gai', 'muc_2sd', 'muc_3sd', 'danh_gia'])

        add(agg['thua_can_cn_tuoi'], s['thua_can_cn_tuoi'], ['tong', 'trai', 'gai'])
        add(agg['beo_phi_cn_tuoi'], s['beo_phi_cn_tuoi'], ['tong', 'trai', 'gai'])
        add(agg['thua_can_beo_phi'], s['thua_can_beo_phi'], ['tong', 'trai', 'gai', 'muc_2sd', 'muc_3sd'])

    total = agg['tong_so_tre']['tong']
    len_can = agg['tre_duoc_can']['tong']
    len_do = agg['tre_duoc_do']['tong']
    # Mẫu số CN/CC = tổng số trẻ được đánh giá CN/CC (đã gộp); fallback len_can nếu thiếu
    len_cn_cc = agg['sdd_cn_cc'].get('danh_gia', 0)

    agg['tre_duoc_can']['ty_le'] = round(len_can / total * 100, 2) if total > 0 else 0
    agg['tre_duoc_do']['ty_le'] = round(len_do / total * 100, 2) if total > 0 else 0

    agg['sdd_cn_tuoi']['ty_le'] = round(agg['sdd_cn_tuoi']['tong'] / len_can * 100, 2) if len_can > 0 else 0
    agg['sdd_cc_tuoi']['ty_le'] = round(agg['sdd_cc_tuoi']['tong'] / len_do * 100, 2) if len_do > 0 else 0
    denom_cn_cc = len_cn_cc if len_cn_cc > 0 else len_can
    agg['sdd_cn_cc']['ty_le'] = round(agg['sdd_cn_cc']['tong'] / denom_cn_cc * 100, 2) if denom_cn_cc > 0 else 0

    return agg


def export_summary_table_to_excel(summary: dict, excel_path: str, sheet_name: str = 'Summary') -> None:
    """Ghi bảng tổng hợp ra Excel (tạo file mới hoặc ghi đè)."""
    df_table = summary_table(summary)
    df_table.to_excel(excel_path, sheet_name=sheet_name, index=False)


def write_column_to_excel(
    df: pd.DataFrame,
    df_column: str,
    excel_path: str,
    excel_column: str,
    start_row: int = 7,
    sheet_name: str = 'Sheet1'
) -> None:
    """
    Ghi đè 1 cột từ DataFrame vào 1 cột trong file Excel.
    Chỉ ghi đè phần nội dung, giữ nguyên header và format.
    
    Args:
        df: DataFrame chứa dữ liệu cần ghi
        df_column: Tên cột trong DataFrame cần lấy dữ liệu
        excel_path: Đường dẫn file Excel
        excel_column: Tên cột trong Excel (A, B, C, ... hoặc AA, AB, ...)
        start_row: Hàng bắt đầu ghi (mặc định là 7)
        sheet_name: Tên sheet (mặc định là 'Sheet1')
    """
    # Load workbook - giữ nguyên format
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    
    if ws is None:
        wb.close()
        raise ValueError("Not found sheet in Excel file.")
    
    # Lấy dữ liệu từ DataFrame
    data = df[df_column].tolist()
    
    # Ghi từng giá trị vào Excel - chỉ thay đổi value, giữ nguyên format của cell
    for i, value in enumerate(data):
        row = start_row + i
        cell = ws[f"{excel_column}{row}"]
        
        # Xử lý giá trị NaN/None - chỉ ghi value, không thay đổi style
        if pd.isna(value):
            cell.value = None
        else:
            cell.value = value
    
    # Lưu file
    wb.save(excel_path)
    wb.close()
    
    print(f"Đã ghi {len(data)} giá trị từ cột '{df_column}' vào cột {excel_column} (từ hàng {start_row})")
