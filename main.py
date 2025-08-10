import openpyxl
from random import randrange
from tkinter import *
from tkinter.ttk import *
from tkinter import messagebox
from tkinter.filedialog import Open
from tkinter import scrolledtext
from threading import Thread
import tkinter


def getValue1(pos):
    global load1
    global sheet1
    return sheet1[pos].value

def getValue2(pos):
    global load2
    global sheet2
    return sheet2[pos].value

def getValue3(pos):
    global load3
    global sheet3
    return sheet3[pos].value

def getValue4(pos):
    global load4
    global sheet4
    return sheet4[pos].value

def update_Value(pos, value):
    sheet0[pos].value = value
    

class class_thang():
    def __init__(self, x, y):
        self.x = x
        self.y = y

pos_thang_nam = [class_thang('G', 5), class_thang('F', 5), class_thang('E', 5), class_thang('D', 5), class_thang('C', 5)
             , class_thang('N', 9), class_thang('M', 9), class_thang('L', 9), class_thang('K', 9), class_thang('J', 9)
             , class_thang('I', 9), class_thang('H', 9), class_thang('G', 9), class_thang('F', 9), class_thang('E', 9)
             , class_thang('D', 9), class_thang('C', 9), class_thang('N', 13), class_thang('M', 13), class_thang('L', 13)
             , class_thang('K', 13), class_thang('J', 13), class_thang('I', 13), class_thang('H', 13), class_thang('G', 13)
             , class_thang('F', 13), class_thang('E', 13), class_thang('D', 13), class_thang('C', 13), class_thang('N', 17)
             , class_thang('M', 17), class_thang('L', 17), class_thang('K', 17), class_thang('J', 17), class_thang('I', 17)
             , class_thang('H', 17), class_thang('G', 17), class_thang('F', 17), class_thang('E', 17), class_thang('D', 17)
             , class_thang('C', 17), class_thang('N', 21), class_thang('M', 21), class_thang('L', 21), class_thang('K', 21)
             , class_thang('J', 21), class_thang('I', 21), class_thang('H', 21), class_thang('G', 21), class_thang('F', 21)
             , class_thang('E', 21), class_thang('D', 21), class_thang('C', 21), class_thang('N', 25), class_thang('M', 25)
             , class_thang('L', 25), class_thang('K', 25), class_thang('J', 25), class_thang('I', 25), class_thang('H', 25)]

pos_thang_nu = [class_thang('G', 36), class_thang('F', 36), class_thang('E', 36), class_thang('D', 36), class_thang('C', 36)
             , class_thang('N', 40), class_thang('M', 40), class_thang('L', 40), class_thang('K', 40), class_thang('J', 40)
             , class_thang('I', 40), class_thang('H', 40), class_thang('G', 40), class_thang('F', 40), class_thang('E', 40)
             , class_thang('D', 40), class_thang('C', 40), class_thang('N', 44), class_thang('M', 44), class_thang('L', 44)
             , class_thang('K', 44), class_thang('J', 44), class_thang('I', 44), class_thang('H', 44), class_thang('G', 44)
             , class_thang('F', 44), class_thang('E', 44), class_thang('D', 44), class_thang('C', 44), class_thang('N', 48)
             , class_thang('M', 48), class_thang('L', 48), class_thang('K', 48), class_thang('J', 48), class_thang('I', 48)
             , class_thang('H', 48), class_thang('G', 48), class_thang('F', 48), class_thang('E', 48), class_thang('D', 48)
             , class_thang('C', 48), class_thang('N', 52), class_thang('M', 52), class_thang('L', 52), class_thang('K', 52)
             , class_thang('J', 52), class_thang('I', 52), class_thang('H', 52), class_thang('G', 52), class_thang('F', 52)
             , class_thang('E', 52), class_thang('D', 52), class_thang('C', 52), class_thang('N', 56), class_thang('M', 56)
             , class_thang('L', 56), class_thang('K', 56), class_thang('J', 56), class_thang('I', 56), class_thang('H', 56)]



def thongKe_SDD():
    soTre = [0, 0]
    phi2T = 0
    phi5T = [0, 0]
    thang = [0, 0, 0, 0, 0, 0]
    duocCanNamNu = [0, 0]
    duocCanThang = [0, 0, 0, 0, 0, 0]
    kDuocCanNamNu = [0, 0]
    kDuocCanThang = [0, 0, 0, 0, 0, 0]
    btCNNamNu = [0, 0]
    btCNThang = [0, 0, 0, 0, 0, 0]
    duocDoNamNu = [0, 0]
    duocDoThang = [0, 0, 0, 0, 0, 0]
    kDuocDoNamNu = [0, 0]
    kDuocDoThang = [0, 0, 0, 0, 0, 0]
    btCCNamNu = [0, 0]
    sddCNNamNu = [0, 0]
    sddCNThang = [0, 0, 0, 0, 0, 0]
    sddCCNamNu = [0, 0]
    sddCCThang = [0, 0, 0, 0, 0, 0]

    so_tre = 1000
    for i in range(1, so_tre + 1):
        THANG = getValue1('F' + str(7 + i))

        if type(THANG) != int:
            break

        if getValue1('C' + str(7 + i)) == 'x':
            NAM = True
        else:
            NAM = False

        CN = getValue1('H' + str(7 + i))
        CC = getValue1('K' + str(7 + i))
        BTCN = getValue1('J' + str(7 + i))
        SDDCN = getValue1('I' + str(7 + i))
        PHI = getValue1('O' + str(7 + i))
        BTCC = getValue1('M' + str(7 + i))
        SDDCC = getValue1('L' + str(7 + i))


        if THANG <= 6:
            thang[0] += 1
            if BTCN == 'BT':
                btCNThang[0] += 1
            elif SDDCN == 'SDD':
                sddCNThang[0] += 1
            elif PHI == 'PHÌ':
                phi2T += 1
            if SDDCC == 'SDD':
                sddCCThang[0] += 1
            if type(CN) != int and type(CN) != float:
                kDuocCanThang[0] += 1
            else:
                duocCanThang[0] += 1
            if type(CC) != int and type(CC) != float:
                kDuocDoThang[0] += 1
            else:
                duocDoThang[0] += 1
        elif THANG <= 12:
            thang[1] += 1
            if BTCN == 'BT':
                btCNThang[1] += 1
            elif SDDCN == 'SDD':
                sddCNThang[1] += 1
            elif PHI == 'PHÌ':
                phi2T += 1
            if SDDCC == 'SDD':
                sddCCThang[1] += 1
            if type(CN) != int and type(CN) != float:
                kDuocCanThang[1] += 1
            else:
                duocCanThang[1] += 1
            if type(CC) != int and type(CC) != float:
                kDuocDoThang[1] += 1
            else:
                duocDoThang[1] += 1
        elif THANG <= 24:
            thang[2] += 1
            if BTCN == 'BT':
                btCNThang[2] += 1
            elif SDDCN == 'SDD':
                sddCNThang[2] += 1
            elif PHI == 'PHÌ':
                phi2T += 1
            if SDDCC == 'SDD':
                sddCCThang[2] += 1
            if type(CN) != int and type(CN) != float:
                kDuocCanThang[2] += 1
            else:
                duocCanThang[2] += 1
            if type(CC) != int and type(CC) != float:
                kDuocDoThang[2] += 1
            else:
                duocDoThang[2] += 1
        elif THANG <= 36:
            thang[3] += 1
            if BTCN == 'BT':
                btCNThang[3] += 1
            elif SDDCN == 'SDD':
                sddCNThang[3] += 1
            if SDDCC == 'SDD':
                sddCCThang[3] += 1
            if type(CN) != int and type(CN) != float:
                kDuocCanThang[3] += 1
            else:
                duocCanThang[3] += 1
            if type(CC) != int and type(CC) != float:
                kDuocDoThang[3] += 1
            else:
                duocDoThang[3] += 1
        elif THANG <= 48:
            thang[4] += 1
            if BTCN == 'BT':
                btCNThang[4] += 1
            elif SDDCN == 'SDD':
                sddCNThang[4] += 1
            if SDDCC == 'SDD':
                sddCCThang[4] += 1
            if type(CN) != int and type(CN) != float:
                kDuocCanThang[4] += 1
            else:
                duocCanThang[4] += 1
            if type(CC) != int and type(CC) != float:
                kDuocDoThang[4] += 1
            else:
                duocDoThang[4] += 1
        elif THANG <= 60:
            thang[5] += 1
            if BTCN == 'BT':
                btCNThang[5] += 1
            elif SDDCN == 'SDD':
                sddCNThang[5] += 1
            if SDDCC == 'SDD':
                sddCCThang[5] += 1
            if type(CN) != int and type(CN) != float:
                kDuocCanThang[5] += 1
            else:
                duocCanThang[5] += 1
            if type(CC) != int and type(CC) != float:
                kDuocDoThang[5] += 1
            else:
                duocDoThang[5] += 1

        if NAM:
            soTre[0] += 1
            if PHI == 'PHÌ':
                phi5T[0] += 1
            elif BTCN == 'BT':
                btCNNamNu[0] += 1
            elif SDDCN == 'SDD':
                sddCNNamNu[0] += 1
            if BTCC == 'BT':
                btCCNamNu[0] += 1
            elif SDDCC == 'SDD':
                sddCCNamNu[0] += 1
            if type(CN) != int and type(CN) != float:
                kDuocCanNamNu[0] += 1
            else:
                duocCanNamNu[0] += 1
            if type(CC) != int and type(CC) != float:
                kDuocDoNamNu[0] += 1
            else:
                duocDoNamNu[0] += 1
        else:
            soTre[1] += 1
            if PHI == 'PHÌ':
                phi5T[1] += 1
            elif BTCN == 'BT':
                btCNNamNu[1] += 1
            elif SDDCN == 'SDD':
                sddCNNamNu[1] += 1
            if BTCC == 'BT':
                btCCNamNu[1] += 1
            elif SDDCC == 'SDD':
                sddCCNamNu[1] += 1
            if type(CN) != int and type(CN) != float:
                kDuocCanNamNu[1] += 1
            else:
                duocCanNamNu[1] += 1
            if type(CC) != int and type(CC) != float:
                kDuocDoNamNu[1] += 1
            else:
                duocDoNamNu[1] += 1

    txtThongKe = '\n2 tuổi: \n' \
                 '\t trẻ dưới 2 tuổi: ' + str(thang[0] + thang[1] + thang[2]) + '\n'\
                 '\t trẻ được cân 2 tuổi: ' + str(duocCanThang[0] + duocCanThang[1] + duocCanThang[2]) + '\n'\
                 '\t trẻ không được cân 2 tuổi: ' + str(kDuocCanThang[0] + kDuocCanThang[1] + kDuocCanThang[2]) + '\n'\
                 '\t trẻ SDD cân nặng 2 tuổi: ' + str(sddCNThang[0] + sddCNThang[1] + sddCNThang[2]) + '\n'\
                 '\t trẻ BT cân nặng 2 tuổi: ' + str(btCNThang[0] + btCNThang[1] + btCNThang[2]) + '\n'\
                 '\t trẻ phì 2 tuổi: ' + str(phi2T) + '\n'\
                 '' \
                 '\n5 tuổi: ' \
                 '\n\t trẻ dưới 5 tuổi: ' + str(soTre[0] + soTre[1]) + '\n'\
                 '\t\t trẻ nam 5 tuổi: ' + str(soTre[0]) + '\n'\
                 '\t\t trẻ nữ 5 tuổi: ' + str(soTre[1]) + '\n'\
                 '\t\t trẻ 0-6 tháng: ' + str(thang[0]) + '\n'\
                 '\t\t trẻ 7-12 tháng: ' + str(thang[1]) + '\n'\
                 '\t\t trẻ 13-24 tháng: ' + str(thang[2]) + '\n'\
                 '\t\t trẻ 25-36 tháng: ' + str(thang[3]) + '\n'\
                 '\t\t trẻ 37-48 tháng: ' + str(thang[4]) + '\n'\
                 '\t\t trẻ 49-60 tháng: ' + str(thang[5]) + '\n'\
                 '' \
                 '\n\t trẻ được cân 5 tuổi: ' + str(duocCanNamNu[0] + duocCanNamNu[1]) + '\n'\
                 '\t\t Được cân nam: ' + str(duocCanNamNu[0]) + '\n'\
                 '\t\t Được cân nữ: ' + str(duocCanNamNu[1]) + '\n'\
                 '\t\t Được cân 6: ' + str(duocCanThang[0]) + '\n'\
                 '\t\t Được cân 12: ' + str(duocCanThang[1]) + '\n'\
                 '\t\t Được cân 24: ' + str(duocCanThang[2]) + '\n'\
                 '\t\t Được cân 36: ' + str(duocCanThang[3]) + '\n'\
                 '\t\t Được cân 48: ' + str(duocCanThang[4]) + '\n'\
                 '\t\t Được cân 60: ' + str(duocCanThang[5]) + '\n'\
                 '' \
                 '\n\t trẻ không được cân 5 tuổi: ' + str(kDuocCanNamNu[0] + kDuocCanNamNu[1]) + '\n'\
                 '\t\t Không được cân nam: ' + str(kDuocCanNamNu[0]) + '\n'\
                 '\t\t Không được cân nữ: ' + str(kDuocCanNamNu[1]) + '\n'\
                 '\t\t Không được cân 6: ' + str(kDuocCanThang[0]) + '\n'\
                 '\t\t Không được cân 12: ' + str(kDuocCanThang[1]) + '\n'\
                 '\t\t Không được cân 24: ' + str(kDuocCanThang[2]) + '\n'\
                 '\t\t Không được cân 36: ' + str(kDuocCanThang[3]) + '\n'\
                 '\t\t Không được cân 48: ' + str(kDuocCanThang[4]) + '\n'\
                 '\t\t Không được cân 60: ' + str(kDuocCanThang[5]) + '\n'\
                 '' \
                 '\n\t trẻ BT cân nặng 5 tuổi: ' + str(btCNNamNu[0] + btCNNamNu[1]) + '\n'\
                 '\t\t BT căn nặng 5 tuổi nam: ' + str(btCNNamNu[0]) + '\n'\
                 '\t\t BT căn nặng 5 tuổi nữ: ' + str(btCNNamNu[1]) + '\n'\
                 '' \
                 '\n\t trẻ SDD cân nặng 5 tuổi: ' + str(sddCNNamNu[0] + sddCNNamNu[1]) + '\n'\
                 '\t\t SDD cân nặng 5 tuổi nam: ' + str(sddCNNamNu[0]) + '\n'\
                 '\t\t SDD cân nặng 5 tuổi nữ: ' + str(sddCNNamNu[1]) + '\n'\
                 '\t\t SDD cân nặng 6 tháng: ' + str(sddCNThang[0]) + '\n'\
                 '\t\t SDD cân nặng 12 tháng: ' + str(sddCNThang[1]) + '\n'\
                 '\t\t SDD cân nặng 24 tháng: ' + str(sddCNThang[2]) + '\n'\
                 '\t\t SDD cân nặng 36 tháng: ' + str(sddCNThang[3]) + '\n'\
                 '\t\t SDD cân nặng 48 tháng: ' + str(sddCNThang[4]) + '\n'\
                 '\t\t SDD cân nặng 60 tháng: ' + str(sddCNThang[5]) + '\n'\
                 '' \
                 '\n\t trẻ được đo 5 tuổi: ' + str(duocDoNamNu[0] + duocDoNamNu[1]) + '\n'\
                 '\t\t Được đo nam: ' + str(duocDoNamNu[0]) + '\n'\
                 '\t\t Được đo nữ: ' + str(duocDoNamNu[1]) + '\n'\
                 '\t\t Được đo 6: ' + str(duocDoThang[0]) + '\n'\
                 '\t\t Được đo 12: ' + str(duocDoThang[1]) + '\n'\
                 '\t\t Được đo 24: ' + str(duocDoThang[2]) + '\n'\
                 '\t\t Được đo 36: ' + str(duocDoThang[3]) + '\n'\
                 '\t\t Được đo 48: ' + str(duocDoThang[4]) + '\n'\
                 '\t\t Được đo 60: ' + str(duocDoThang[5]) + '\n'\
                 '' \
                 '\n\t trẻ không được đo 5 tuổi: ' + str(kDuocDoNamNu[0] + kDuocDoNamNu[1]) + '\n'\
                 '\t\t Không được cân nam: ' + str(kDuocDoNamNu[0]) + '\n'\
                 '\t\t Không được đo nữ: ' + str(kDuocDoNamNu[1]) + '\n'\
                 '\t\t Không được đo 6: ' + str(kDuocDoThang[0]) + '\n'\
                 '\t\t Không được đo 12: ' + str(kDuocDoThang[1]) + '\n'\
                 '\t\t Không được đo 24: ' + str(kDuocDoThang[2]) + '\n'\
                 '\t\t Không được đo 36: ' + str(kDuocDoThang[3]) + '\n'\
                 '\t\t Không được đo 48: ' + str(kDuocDoThang[4]) + '\n'\
                 '\t\t Không được đo 60: ' + str(kDuocDoThang[5]) + '\n'\
                 '' \
                 '\n\t trẻ BT chiều cao 5 tuổi: ' + str(btCCNamNu[0] + btCCNamNu[1]) + '\n'\
                 '\t\t BT chiều cao 5 tuổi nam: ' + str(btCCNamNu[0]) + '\n'\
                 '\t\t BT chiều cao 5 tuổi nữ: ' + str(btCCNamNu[1]) + '\n'\
                 '' \
                 '\n\t trẻ SDD chiều cao: ' + str(sddCCNamNu[0] + sddCCNamNu[1]) + '\n'\
                 '\t\t SDD chiều cao 5 tuổi nam: ' + str(sddCCNamNu[0]) + '\n'\
                 '\t\t SDD chiều cao 5 tuổi nữ: ' + str(sddCCNamNu[1]) + '\n'\
                 '\t\t SDD chiều cao 6 tháng: ' + str(sddCCThang[0]) + '\n'\
                 '\t\t SDD chiều cao 12 tháng: ' + str(sddCCThang[1]) + '\n'\
                 '\t\t SDD chiều cao 24 tháng: ' + str(sddCCThang[2]) + '\n'\
                 '\t\t SDD chiều cao 36 tháng: ' + str(sddCCThang[3]) + '\n'\
                 '\t\t SDD chiều cao 48 tháng: ' + str(sddCCThang[4]) + '\n'\
                 '\t\t SDD chiều cao 60 tháng: ' + str(sddCCThang[5]) + '\n'\
                 '' \
                 '\n\t trẻ Phì 5 tuổi: ' + str(phi5T[0] + phi5T[1]) + '\n'\
                 '\t\t Phì 5 tuổi nam: ' + str(phi5T[0]) + '\n'\
                 '\t\t Phì 5 tuổi nữ: ' + str(phi5T[1]) + '\n'\

    return txtThongKe


def thongKe_CNCC():
    btNamNu = [0, 0]
    btThang = [0, 0, 0, 0, 0, 0]
    sdd1NamNu = [0, 0]
    sdd1Thang = [0, 0, 0, 0, 0, 0]
    sdd2NamNu = [0, 0]
    sdd2Thang = [0, 0, 0, 0, 0, 0]

    so_tre = 1000
    for i in range(1, so_tre + 1):

        THANG = getValue1('F' + str(7 + i))

        if type(THANG) != int:
            break

        if getValue1('C' + str(7 + i)) == 'x':
            NAM = True
        else:
            NAM = False

        CNCC = getValue1('N' + str(7 + i))
        if type(CNCC) != str:
            continue

        if THANG <= 6:
            if CNCC == 'BT':
                btThang[0] += 1
            elif CNCC == 'Độ I':
                sdd1Thang[0] += 1
            elif CNCC == 'Độ II':
                sdd2Thang[0] += 1
        elif THANG <= 12:
            if CNCC == 'BT':
                btThang[1] += 1
            elif CNCC == 'Độ I':
                sdd1Thang[1] += 1
            elif CNCC == 'Độ II':
                sdd2Thang[1] += 1
        elif THANG <= 24:
            if CNCC == 'BT':
                btThang[2] += 1
            elif CNCC == 'Độ I':
                sdd1Thang[2] += 1
            elif CNCC == 'Độ II':
                sdd2Thang[2] += 1
        elif THANG <= 36:
            if CNCC == 'BT':
                btThang[3] += 1
            elif CNCC == 'Độ I':
                sdd1Thang[3] += 1
            elif CNCC == 'Độ II':
                sdd2Thang[3] += 1
        elif THANG <= 48:
            if CNCC == 'BT':
                btThang[4] += 1
            elif CNCC == 'Độ I':
                sdd1Thang[4] += 1
            elif CNCC == 'Độ II':
                sdd2Thang[4] += 1
        elif THANG <= 60:
            if CNCC == 'BT':
                btThang[5] += 1
            elif CNCC == 'Độ I':
                sdd1Thang[5] += 1
            elif CNCC == 'Độ II':
                sdd2Thang[5] += 1

        if NAM:
            if CNCC == 'BT':
                btNamNu[0] += 1
            elif CNCC == 'Độ I':
                sdd1NamNu[0] += 1
            elif CNCC == 'Độ II':
                sdd2NamNu[0] += 1
        else:
            if CNCC == 'BT':
                btNamNu[1] += 1
            elif CNCC == 'Độ I':
                sdd1NamNu[1] += 1
            elif CNCC == 'Độ II':
                sdd2NamNu[1] += 1

    txtThongKe = '\t trẻ BT CN/CC: ' + str(btNamNu[0] + btNamNu[1]) + '\n'\
                 '\t\t BT CN/CC nam: ' + str(btNamNu[0]) + '\n'\
                 '\t\t BT CN/CC nữ: ' + str(btNamNu[1]) + '\n'\
                 '\t\t BT CN/CC 6 tháng: ' + str(btThang[0]) + '\n' \
                 '\t\t BT CN/CC 12 tháng: ' + str(btThang[1]) + '\n' \
                 '\t\t BT CN/CC 24 tháng: ' + str(btThang[2]) + '\n'\
                 '\t\t BT CN/CC 36 tháng: ' + str(btThang[3]) + '\n'\
                 '\t\t BT CN/CC 48 tháng: ' + str(btThang[4]) + '\n' \
                 '\t\t BT CN/CC 60 tháng: ' + str(btThang[5]) + '\n' \
                 '' \
                 '\n\t trẻ SDD CN/CC độ I: ' + str(sdd1NamNu[0] + sdd1NamNu[1]) + '\n'\
                 '\t\t SDD CN/CC độ I nam: ' + str(sdd1NamNu[0]) + '\n'\
                 '\t\t SDD CN/CC độ I nữ: ' + str(sdd1NamNu[1]) + '\n'\
                 '\t\t SDD CN/CC độ I 6 tháng: ' + str(sdd1Thang[0]) + '\n'\
                 '\t\t SDD CN/CC độ I 12 tháng: ' + str(sdd1Thang[1]) + '\n'\
                 '\t\t SDD CN/CC độ I 24 tháng: ' + str(sdd1Thang[2]) + '\n'\
                 '\t\t SDD CN/CC độ I 36 tháng: ' + str(sdd1Thang[3]) + '\n'\
                 '\t\t SDD CN/CC độ I 48 tháng: ' + str(sdd1Thang[4]) + '\n'\
                 '\t\t SDD CN/CC độ I 60 tháng: ' + str(sdd1Thang[5]) + '\n' \
                 '' \
                 '\n\t trẻ SDD CN/CC độ II: ' + str(sdd2NamNu[0] + sdd2NamNu[1]) + '\n' \
                 '\t\t SDD CN/CC độ II nam: ' + str(sdd2NamNu[0]) + '\n' \
                 '\t\t SDD CN/CC độ II nữ: ' + str(sdd2NamNu[1]) + '\n' \
                 '\t\t SDD CN/CC độ II 6 tháng: ' + str(sdd2Thang[0]) + '\n' \
                 '\t\t SDD CN/CC độ II 12 tháng: ' + str(sdd2Thang[1]) + '\n' \
                 '\t\t SDD CN/CC độ II 24 tháng: ' + str(sdd2Thang[2]) + '\n' \
                 '\t\t SDD CN/CC độ II 36 tháng: ' + str(sdd2Thang[3]) + '\n' \
                 '\t\t SDD CN/CC độ II 48 tháng: ' + str(sdd2Thang[4]) + '\n' \
                 '\t\t SDD CN/CC độ II 60 tháng: ' + str(sdd2Thang[5]) + '\n' \

    return txtThongKe


def chamSDD():
    #so_tre = int(input('Nhap so tre: '))
    so_tre = 1000
    for i in range(1, so_tre + 1):

        if getValue1('C' + str(7 + i)) == 'x':
            nam = True
            nu = False
        else:
            nu = True
            nam = False

        thang = getValue1('F' + str(7 + i))
        if type(thang) != int:
            if type(getValue1('B' + str(7 + i))) != str and type(getValue1('G' + str(7 + i))) != str:
                break
            else:
                continue

        can_Nang = getValue1('H' + str(7 + i))

        chieu_Cao = getValue1('K' + str(7 + i))



        if nam:
            if can_Nang == 'vắng' or (type(can_Nang) != int and type(can_Nang) != float):
                if can_Nang != 'vắng':
                    print("xem lại hàng ", 7 + i)
            else:
                if can_Nang < getValue2(pos_thang_nam[thang - 1].x + str(pos_thang_nam[thang - 1].y + 1)):
                    update_Value('I' + str(7 + i), 'SDD')
                    print(getValue1('B' + str(7 + i)), 'SDD(cân_nặng)')
                elif can_Nang > getValue2(pos_thang_nam[thang - 1].x + str(pos_thang_nam[thang - 1].y + 2)):
                    update_Value('O' + str(7 + i), 'PHÌ')
                    print(getValue1('B' + str(7 + i)), 'PHÌ(cân_nặng)')
                else:
                    update_Value('J' + str(7 + i), 'BT')

            if chieu_Cao == 'vắng' or (type(chieu_Cao) != int and type(chieu_Cao) != float):
                if chieu_Cao != 'vắng':
                    print("xem lại hàng ", 7 + i)
                continue
            if chieu_Cao < getValue2(pos_thang_nam[thang - 1].x + str(pos_thang_nam[thang - 1].y - 1)):
                update_Value('L' + str(7 + i), 'SDD')
                print(getValue1('B' + str(7 + i)), 'SDD(chiều_cao)')
            else:
                update_Value('M' + str(7 + i), 'BT')

        elif nu:
            if can_Nang == 'vắng' or (type(can_Nang) != int and type(can_Nang) != float):
                if can_Nang != 'vắng':
                    print("xem lại hàng ", 7 + i)
            else:
                if can_Nang < getValue2(pos_thang_nu[thang - 1].x + str(pos_thang_nu[thang - 1].y + 1)):
                    update_Value('I' + str(7 + i), 'SDD')
                    print(getValue1('B' + str(7 + i)), 'SDD(cân_nặng)')
                elif can_Nang > getValue2(pos_thang_nu[thang - 1].x + str(pos_thang_nu[thang - 1].y + 2)):
                    update_Value('O' + str(7 + i), 'PHÌ')
                    print(getValue1('B' + str(7 + i)), 'PHÌ(cân_nặng)')
                else:
                    update_Value('J' + str(7 + i), 'BT')

            if chieu_Cao == 'vắng' or (type(chieu_Cao) != int and type(chieu_Cao) != float):
                if chieu_Cao != 'vắng':
                    print("xem lại hàng ", 7 + i)
                continue
            if chieu_Cao < getValue2(pos_thang_nu[thang - 1].x + str(pos_thang_nu[thang - 1].y - 1)):
                update_Value('L' + str(7 + i), 'SDD')
                print(getValue1('B' + str(7 + i)), 'SDD(chiều_cao)')
            else:
                update_Value('M' + str(7 + i), 'BT')
    
    load0.save(filename1)
    load0.close()

    messagebox.showinfo('', 'Hoàn thành chấm SDD')


def chamCNCC():
    abcdnam = 'HI'
    abcdnu = 'KL'
    abcnam = 'BC'
    abcnu = 'EF'
    check = FALSE
    # so_tre = int(input('Nhap so tre: '))
    so_tre = 1000

    for i in range(1, so_tre + 1):

        if getValue1('C' + str(7 + i)) == 'x':
            nam = True
            nu = False
        else:
            nu = True
            nam = False

        thang = getValue1('F' + str(7 + i))
        if type(thang) != int:
            if type(getValue1('B' + str(7 + i))) != str and type(getValue1('G' + str(7 + i))) != str:
                break
            else:
                continue

        can_Nang = getValue1('H' + str(7 + i))
        if can_Nang == 'vắng' or (type(can_Nang) != int and type(can_Nang) != float):
            if can_Nang != 'vắng':
                print("xem lại hàng ", 7 + i)
            continue

        chieu_Cao = getValue1('K' + str(7 + i))
        if chieu_Cao == 'vắng' or (type(chieu_Cao) != int and type(chieu_Cao) != float):
            if chieu_Cao != 'vắng':
                print("xem lại hàng ", 7 + i)
            continue


        j = int(chieu_Cao) - 38
        check = FALSE

        if thang <= 24:
            if nam:
                for c in abcnam:
                    if can_Nang < getValue4(c + str(j)):
                        update_Value('N' + str(i + 7), str(getValue4(c + str(6))))
                        check = TRUE
                        break

            else:
                for c in abcnu:
                    if can_Nang < getValue4(c + str(j)):
                        update_Value('N' + str(i + 7), str(getValue4(c + str(6))))
                        check = TRUE
                        break
        else:
            if nam:
                for c in abcdnam:
                    if can_Nang < getValue4(c + str(j)):
                        update_Value('N' + str(i + 7), str(getValue4(c + str(6))))
                        check = TRUE
                        break
            else:
                for c in abcdnu:
                    if can_Nang < getValue4(c + str(j)):
                        update_Value('N' + str(i + 7), str(getValue4(c + str(6))))
                        check = TRUE
                        break


        if check == FALSE:
            update_Value('N' + str(i + 7), 'BT')

    load0.save(filename1)
    load0.close()

    messagebox.showinfo('', 'Hoàn thành chấm CN/CC')


def chamSDDII():
    #so_tre = int(input('Nhập số trẻ: '))
    so_tre = 1000
    for i in range(1, so_tre + 1):

        if getValue1('C' + str(5 + i)) == 'x':
            nam = True
            nu = False
        else:
            nu = True
            nam = False

        thang = getValue1('F' + str(5 + i))
        if type(thang) != int:
            break
        chieu_Cao = getValue1('L' + str(5 + i))
        can_Nang = getValue1('H' + str(5 + i))

        if nam:
            if can_Nang < getValue3(pos_thang_nam[thang - 1].x + str(pos_thang_nam[thang - 1].y + 1)):
                update_Value('K' + str(5 + i), 'x')
                print(getValue1('B' + str(5 + i)), 'SDD độ II(cân nặng)')
            elif getValue1('I' + str(5 + i)) == 'SDD':
                update_Value('J' + str(5 + i), 'x')
            if chieu_Cao < getValue3(pos_thang_nam[thang - 1].x + str(pos_thang_nam[thang - 1].y - 1)):
                update_Value('O' + str(5 + i), 'x')
                print(getValue1('B' + str(5 + i)), 'SDD độ II(chiều cao)')
            elif getValue1('M' + str(5 + i)) == 'SDD':
                update_Value('N' + str(5 + i), 'x')

        elif nu:
            if can_Nang < getValue3(pos_thang_nu[thang - 1].x + str(pos_thang_nu[thang - 1].y + 1)):
                update_Value('K' + str(5 + i), 'x')
                print(getValue1('B' + str(5 + i)), 'SDD độ II(cân nặng)')
            elif getValue1('I' + str(5 + i)) == 'SDD':
                update_Value('J' + str(5 + i), 'x')
            if chieu_Cao < getValue3(pos_thang_nu[thang - 1].x + str(pos_thang_nu[thang - 1].y - 1)):
                update_Value('O' + str(5 + i), 'x')
                print(getValue1('B' + str(5 + i)), 'SDD độ II(chiều cao)')
            elif getValue1('M' + str(5 + i)) == 'SDD':
                update_Value('N' + str(5 + i), 'x')

    load0.save(filename1)
    load0.close()

    messagebox.showinfo('', 'Hoàn thành chấm SDD độ II')


def randomCNCC():
    so_tre = 1000
    for i in range(1, so_tre + 1):

        if getValue1('C' + str(7 + i)) == 'x':
            nam = True
            nu = False
        else:
            nu = True
            nam = False

        thang = getValue1('F' + str(7 + i))
        if type(thang) != int:
            if type(getValue1('B' + str(7 + i))) != str and type(getValue1('G' + str(7 + i))) != str:
                break
            else:
                continue

        can_Nang = getValue1('H' + str(7 + i))
        if type(can_Nang) == str:
            continue

        chieu_Cao = getValue1('K' + str(7 + i))
        if type(chieu_Cao) == str:
            continue


        if nam:
            can_Nang_phong = getValue2(pos_thang_nam[thang - 1].x + str(pos_thang_nam[thang - 1].y + 1)) + \
                             randrange(1, 6, 1)/10

            chieu_Cao_phong = getValue2(pos_thang_nam[thang - 1].x + str(pos_thang_nam[thang - 1].y - 1)) + \
                              randrange(1, 11, 1)/10.0

            if (type(can_Nang) != int and type(can_Nang) != float):
                update_Value('H' + str(7 + i), can_Nang_phong)

            if (type(chieu_Cao) != int and type(chieu_Cao) != float):
                update_Value('K' + str(7 + i), chieu_Cao_phong)

        if nu:
            can_Nang_phong = getValue2(pos_thang_nu[thang - 1].x + str(pos_thang_nu[thang - 1].y + 1)) + \
                             randrange(1, 6, 1)/10.0

            chieu_Cao_phong = getValue2(pos_thang_nu[thang - 1].x + str(pos_thang_nu[thang - 1].y - 1)) + \
                              randrange(1, 11, 1)/10.0

            if (type(can_Nang) != int and type(can_Nang) != float):
                update_Value('H' + str(7 + i), can_Nang_phong)

            if chieu_Cao == 'vắng' or (type(chieu_Cao) != int and type(chieu_Cao) != float):
                update_Value('K' + str(7 + i), chieu_Cao_phong)

    load0.save(filename1)
    load0.close()
    
    messagebox.showinfo('', 'Hoàn thành điền CN_CC')


def thongKe_window():
    if var_ThongKe_SDD.get():
        txtThongKe = thongKe_SDD()
    elif var_ThongKe_CNCC.get():
        txtThongKe = thongKe_CNCC()

    windowScrolledText = Tk()
    windowScrolledText.geometry('700x500')
    txt = scrolledtext.ScrolledText(windowScrolledText, width=200, height=50)
    txt.insert(INSERT, txtThongKe)
    txt.pack()
    messagebox.showinfo('', 'Hoàn thành thống kê')
    windowScrolledText.mainloop()

def clickNew():
    global filename1
    global load0
    global sheet0

    ftypes = [('Excel file', '*.xlsx'), ('All files', '*')]
    dlg = Open(filetypes=ftypes)
    filename1 = dlg.show()
    lblFile.configure(text=filename1)
    print(filename1)

    load0 = openpyxl.load_workbook(filename1)
    sheet0 = load0['Sheet1']

def thread_mess(text):
    thread = Thread(target=messagebox.showinfo('', text))
    thread.start()

def handleButton1():
    global load1
    global sheet1
    global sheet2
    global load3
    global sheet3
    global load4
    global sheet4


    if filename1 == '':
        messagebox.showinfo('', 'Chưa chọn file!!')
    else:
        load1 = openpyxl.load_workbook(filename1)
        sheet1 = load1['Sheet1']
        if var_RandomCNCC.get():
            load2 = openpyxl.load_workbook('Bang cham kenh danh cho tre SDD .xlsx')
            sheet2 = load2['Sheet1']
            thread_mess('Đang điền CN_CC')
            randomCNCC()

        if var_SDD.get():
            load2 = openpyxl.load_workbook('Bang cham kenh danh cho tre SDD .xlsx')
            sheet2 = load2['Sheet1']
            thread_mess('Đang chấm SDD')
            chamSDD()

        if var_SDDII.get():
            if var_SDD.get() or var_CNCC.get() or var_ThongKe_SDD.get():
                messagebox.showinfo('', 'Không thể thực hiện chấm SDD độ II cùng các thao tác khác')
            else:
                load3 = openpyxl.load_workbook('Bang cham kenh SDD do II.xlsx')
                sheet3 = load3['Sheet1']
                thread_mess('Đang chấm SDD độ II')
                chamSDDII()

        if var_CNCC.get():
            load4 = openpyxl.load_workbook('Bang tra CNCC.xlsx')
            sheet4 = load4['Sheet1']
            thread_mess('Đang chấm CN/CC')
            chamCNCC()

        if var_ThongKe_SDD.get():
            thread_mess('Đang thống kê SDD')
            thongKe_window()

        if var_ThongKe_CNCC.get():
            thread_mess('Đang thống kê CN/CC')
            thongKe_window()


    #messagebox.showinfo('hello world', 'Hoàn thành')
    return

def handleButton2():
    thread =Thread(target=handleButton1)
    thread.start()



window = Tk()
window.title('Nguyễn Lộc')
window.geometry("700x500")

# Thêm menu bar, Lấy đường dẫn
filename1 = ''
menu = Menu(window)
new_item = Menu(menu, tearoff=0)
new_item.add_command(label='New', command=clickNew)
menu.add_cascade(label='File', menu=new_item)
window.config(menu=menu)

# Thêm label
hello = "Xin Chào!!"
lblHello = tkinter.Label(window, text=hello, fg="red", font=("Times new roman", 20))
lblHello.place(x=0, y=0)

lblFile = Label(window, text='', font=("Times new roman", 13))
lblFile.place(x=0, y=40)

# Thêm combobox
'''combo = Combobox(window)
combo['values'] = ("Chấm SDD", "Chấm SDD độ I-II", "Chấm CN/CC", "Thống kê")
combo.current(0)
combo.grid(column=0, row=2)'''

# Thêm button
btnChon = Button(window, text="Chọn", command=handleButton2)
btnChon.place(x=200, y=200)

# Thêm checkbox
var_SDD = BooleanVar()
var_SDDII = BooleanVar()
var_CNCC = BooleanVar()
var_ThongKe_SDD = BooleanVar()
var_ThongKe_CNCC = BooleanVar()
var_RandomCNCC = BooleanVar()

cb_SDD = Checkbutton(window, text="Chấm SDD", variable=var_SDD)
cb_SDDII = Checkbutton(window, text="Chấm SDD độ II", variable=var_SDDII)
cb_ChamCNCC = Checkbutton(window, text="Chấm CN/CC", variable=var_CNCC)
cb_ThongKe_SDD = Checkbutton(window, text="Thống kê SDD", variable=var_ThongKe_SDD)
cb_ThongKe_CNCC = Checkbutton(window, text="Thống kê CN/CC", variable=var_ThongKe_CNCC)
cb_RandomCNCC = Checkbutton(window, text="Điền CN_CC", variable=var_RandomCNCC)

cb_SDD.place(x=50, y=100)
cb_SDDII.place(x=300, y=100)
cb_RandomCNCC.place(x=550, y=100)
cb_ChamCNCC.place(x=50, y=150)
cb_ThongKe_SDD.place(x=300, y=150)
cb_ThongKe_CNCC.place(x=550, y=150)


window.mainloop()

